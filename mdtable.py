#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Program to convert .xlsx to grid_tables for markdown
"""

import sys, signal
sys.dont_write_bytecode = True
signal.signal(signal.SIGINT, signal.SIG_DFL)

import argparse
import openpyxl
import pyperclip
import re

from mods.func_prompt_io import check_exist



# =============== Constant =============== #
BORDER_VERTICAL = "|"
BORDER_HORIZONTAL = "-"
BORDER_CROSS = "+"
BORDER_HEADER = "="
RE_COORDINATE = re.compile(r"^(\D+)(\d+)$")
RE_NUMBER_FORMAT_VALUE = re.compile(r"0.(0+)")



#%% ==============================
class Cell:
	def __init__(self, coordinate, value):
		self._coordinate = coordinate
		self._value = None
		self._merged_cells = []
		self._borders = [None, None, None, None]	# top, right, bottom, and left borders
		self._style = None
		self._size = [None, None]
		self._is_header = False

		self.set_value(value)


	@property
	def value(self):
		if (self._value is None or self._value == "") and len(self._merged_cells) == 0:
			return " "
		return self._value

	@property
	def coordinate(self):
		return self._coordinate

	@property
	def size(self):
		if self._size == [None, None]:
			self.get_size()
		return self._size

	@property
	def width(self):
		if self._size == [None, None]:
			self.get_size()
		return self._size[1]

	@property
	def height(self):
		if self._size == [None, None]:
			self.get_size()
		return self._size[0]

	@property
	def format_value(self):
		if self._style is not None and (isinstance(self._value, int) or isinstance(self._value, float)):
			return self._style.format(self._value)
		else:
			return self._value

	@property
	def is_merged(self):
		if len(self._merged_cells) == 0:
			return False
		else:
			return True

	@property
	def merged_cells(self):
		return self._merged_cells

	@property
	def borders(self):
		if self._borders == [None, None, None, None]:
			self.set_border()
		return self._borders

	@property
	def has_border_top(self):
		if self._borders == [None, None, None, None]:
			self.set_border()
		return self._borders[0]

	@property
	def has_border_right(self):
		if self._borders == [None, None, None, None]:
			self.set_border()
		return self._borders[1]

	@property
	def has_border_bottom(self):
		if self._borders == [None, None, None, None]:
			self.set_border()
		return self._borders[2]

	@property
	def has_border_left(self):
		if self._borders == [None, None, None, None]:
			self.set_border()
		return self._borders[3]

	@property
	def is_header(self):
		return self._is_header


	def get_size(self):
		"""
		Method to get cell size
		"""
		list_value = str(self.format_value).split("\n")
		self._size = [len(list_value), max([len(v) for v in list_value])]
		return self


	def set_value(self, value):
		"""
		Method to set value

		Args:
			value (any): cell value
		"""
		if value is None:
			self._value = ""
		else:
			self._value = value
		return self


	def append_merge_cell(self, obj_cell):
		"""
		Method to append merged cell

		Args:
			obj_cell (Cell object): Cell object
		"""
		self._merged_cells.append(obj_cell)
		return self


	def set_number_format(self, number_format):
		"""
		Method to set style

		Args:
			number_format (str): number format style
		"""
		self._style = "{0:"
		if "," in number_format:
			self._style += ","

		obj_match = RE_NUMBER_FORMAT_VALUE.search(number_format)
		if obj_match:
			self._style += "." + str(len(obj_match.group(1))) + "f"

		self._style += "}"

		return self


	def set_merged_cell(self, list_obj_cell):
		"""
		Method to set list of Cell object

		Args:
			list_obj_cell (list): list of Cell object
		"""
		self._merged_cells = list_obj_cell
		return self


	def set_border(self):
		"""
		Method to make border information from merged cell information
		"""
		self._borders = [True, True, True, True]	# [top, right, bottom, left]
		if len(self._merged_cells) == 0:
			# not merged cell
			return self

		row_idx_self, col_idx_self = convert_coordinate2index(self._coordinate)
		for obj_cell in self._merged_cells:
			row_idx, col_idx = convert_coordinate2index(obj_cell.coordinate)
			if row_idx_self == row_idx and col_idx_self < col_idx:
				self._borders[1] = False
			elif row_idx_self == row_idx and col_idx_self > col_idx:
				self._borders[3] = False
			elif row_idx_self < row_idx and col_idx_self == col_idx:
				self._borders[2] = False
			elif row_idx_self > row_idx and col_idx_self == col_idx:
				self._borders[0] = False
		return self


	def set_header(self, is_header):
		"""
		Method to set header

		Args:
			is_header (bool): header option
		"""
		self._is_header = is_header
		return self



# =============== Function =============== #
def convert_index2coordinate(row, col):
	"""
	Function to convert cell index to coordinates

	Args:
		row (int): row index starting from 1
		col (int): column index starting from 1

	Returns:
		str: cell coordinate
	"""
	col_name = openpyxl.utils.get_column_letter(col)
	return "{}{}".format(col_name, row)


def convert_coordinate2index(coordinate):
	"""
	Function to convert cell coordinate to index

	Args:
		coordinate (str): cell coordinate

	Returns:
		int: row index
		int: column index
	"""
	obj_match = RE_COORDINATE.search(coordinate)
	col_name, row_idx = obj_match.groups()
	col_idx = openpyxl.utils.column_index_from_string(col_name)
	return int(row_idx), int(col_idx)


def get_cells(input_file, sheetname, cell_area):
	# open workbook
	obj_wb = openpyxl.load_workbook(input_file, data_only=True)

	# open worksheet
	obj_ws = None
	if sheetname is not None:
		if sheetname in obj_wb.sheetnames:
			obj_ws = obj_wb[sheetname]
		else:
			sys.stderr.write("ERROR: sheetname `{}` is not found.\n".format(sheetname))
			sys.exit(1)

	else:
		obj_ws = obj_wb.active

	# determine default cell area
	if cell_area == [None, None]:
		cell_area[0] = "{}{}".format(obj_ws.min_column, obj_ws.min_row)
		cell_area[1] = "{}{}".format(obj_ws.max_column, obj_ws.max_row)

	# read cell and make cell objects
	list_cells = {}
	layout_cells = []
	for row in obj_ws["{}:{}".format(*cell_area)]:
		layout_cells.append([])
		for cell in row:
			obj_cell = Cell(cell.coordinate, cell.value)
			obj_cell.set_number_format(cell.number_format)
			obj_cell.set_header(cell.font.b)	# header option
			layout_cells[-1].append(obj_cell)
			list_cells[cell.coordinate] = obj_cell

	# set merged cell information
	merged_cells = [["{}{}".format(openpyxl.utils.get_column_letter(pos[1]), pos[0]) for pos in obj_cell.cells] for obj_cell in obj_ws.merged_cells.ranges]
	for list_coordinate in merged_cells:
		list_obj_cell = [list_cells[coordinate] for coordinate in list_coordinate if coordinate in list_cells.keys()]
		for obj_cell in list_obj_cell:
			partner_obj_cells = set(list_obj_cell) - set([obj_cell])
			# print(obj_cell.coordinate, [v.coordinate for v in partner_obj_cells])
			obj_cell.set_merged_cell(list(partner_obj_cells))
			for obj_partner_cell in partner_obj_cells:
				obj_partner_cell.set_header(obj_cell.is_header)

	# make one header border
	list_border = [all([obj_cell.is_header for obj_cell in row]) for row in layout_cells]
	row_idx_header = None
	if any(list_border):
		row_idx_header = len(list_border) - list_border[-1::-1].index(True) - 1

	for row_idx, row_val in enumerate(layout_cells):
		for obj_cell in row_val:
			if row_idx == row_idx_header:
				obj_cell.set_header(True)

			else:
				obj_cell.set_header(False)

	return layout_cells


def convert_markdown(layout_cells):
	max_row = len(layout_cells)
	max_col = len(layout_cells[0])

	# check width
	list_width = []
	for col_i in range(max_col):
		width = [layout_cells[row_i][col_i].width+2 for row_i in range(max_row)]
		list_width.append(max(width))
	list_format = ["{0:<"+str(v)+"}" for v in list_width]

	# check height
	list_height = []
	for row_i in range(max_row):
		height = [layout_cells[row_i][col_i].height for col_i in range(max_col)]
		list_height.append(max(height))

	contents = []
	for row_i in range(max_row):
		row = [[] for _ in range(list_height[row_i])]
		border_horizontal_top = []
		border_horizontal_bottom = []
		for col_i in range(max_col):
			obj_cell = layout_cells[row_i][col_i]
			if row_i == 0:
				# first row
				if obj_cell.has_border_top:
					# add horizontal top border
					border_horizontal_top.append(BORDER_HORIZONTAL*list_width[col_i])
				else:
					# no border
					border_horizontal_top.append(" "*list_width[col_i])

			if col_i == 0:
				# first column
				# left border
				if obj_cell.has_border_left:
					# add vertical left border
					for r in row:
						r.append(BORDER_VERTICAL)
				else:
					# no border
					for r in row:
						r.append(" ")

			# add value
			values = str(obj_cell.format_value).split("\n")
			if len(values) < len(row):
				values += [""]*(len(row) - len(values))
			values = [list_format[col_i].format(" "+v+" ") for v in values]
			for r, v  in zip(row, values):
				r.append(v)

			# right border
			if obj_cell.has_border_right:
				# add vertical right border
				for r in row:
					r.append(BORDER_VERTICAL)
			else:
				# no border
				for r in row:
					r.append(" ")

			# bottom border
			if obj_cell.has_border_bottom:
				# add horizontal bottom border
				if obj_cell.is_header:
					border_horizontal_bottom.append(":"+BORDER_HEADER*(list_width[col_i]-2)+":")
				else:
					border_horizontal_bottom.append(BORDER_HORIZONTAL*list_width[col_i])
			else:
				# no border
				border_horizontal_bottom.append(" "*list_width[col_i])

		# add top border
		if len(border_horizontal_top) != 0:
			border_horizontal_top = [""] + border_horizontal_top + [""]
			contents.append(BORDER_CROSS.join(border_horizontal_top))
		# add row
		for r in row:
			contents.append("".join(r))

		# add bottom border
		if len(border_horizontal_bottom) != 0:
			border_horizontal_bottom = [""] + border_horizontal_bottom + [""]
			contents.append(BORDER_CROSS.join(border_horizontal_bottom))

	return "\n".join(contents)



# =============== main =============== #
if __name__ == '__main__':
	parser = argparse.ArgumentParser(description="Program to convert .xlsx to grid_tables for markdown", formatter_class=argparse.RawTextHelpFormatter)
	parser.add_argument("-i", dest="INPUT_FILE", metavar="INPUT.xlsx", required=True, help="input .xlsx file")
	parser.add_argument("-s", dest="SHEET_NAME", metavar="SHEET_NAME", help="sheet name for converting table")
	parser.add_argument("-r", dest="CELL_AREA", metavar="CELL_NAME", nargs=2, required=True, default=[None, None], help="Start and End position cells for target square area")
	parser.add_argument("-c", dest="TO_CLIPBOARD", action="store_true", default=False, help="send clipboard")
	args = parser.parse_args()

	check_exist(args.INPUT_FILE, 2)

	layout_cells = get_cells(args.INPUT_FILE, args.SHEET_NAME, args.CELL_AREA)
	str_grid_table = convert_markdown(layout_cells)

	if args.TO_CLIPBOARD:
		pyperclip.copy(str_grid_table+"\n")
	else:
		print(str_grid_table+"\n")
